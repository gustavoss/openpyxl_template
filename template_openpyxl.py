from openpyxl import load_workbook
from openpyxl import styles
import openpyxl

#carrega planilha
wb = load_workbook(filename='template.xlsx', data_only=True)

#conta número de abas
sheets_number = len(wb.worksheets)
print("Quantidade de abas: " + str(sheets_number))

#encontra ultima linha preenchida da primeira aba
last_row_aba1 = wb.worksheets[0].max_row
print("Ultima Linha aba 1: " + str(last_row_aba1))

#encontra ultima coluna preenchida
last_col_aba1 = wb.worksheets[0].max_column
print("Ultima Coluna aba 1: " + str(last_col_aba1))

#ler valor de celula em loop
for i in range(2, last_row_aba1):
    valor_celula = wb.worksheets[0].cell(row=i, column=1).value
    print(valor_celula)

#escreve valor na primeira linha disponível na coluna A
wb.worksheets[0].cell(row=last_row_aba1 + 1, column=1).value = "teste"

#transfere valores da primeira aba para a segunda
last_row_aba2 = wb.worksheets[1].max_row
for i in range(1, last_row_aba1):
        wb.worksheets[1].cell(row=i, column=1).value = wb.worksheets[0].cell(row=i, column=1).value

#cria variavel que recebe uma cor RGB hexadecimal
my_color = openpyxl.styles.colors.Color(rgb='00FF0000')
#cria variavel que adapta a cor my_color para ser aplicada como cor sólida a uma celula
fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_color)

#pinta a segunda célula da primeira coluna da aba1 de vermelho:
wb.worksheets[0].cell(row=2, column=1).fill = fill

#salva planilha
wb.save('template.xlsx')